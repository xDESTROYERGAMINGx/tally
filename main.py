from kivy.config import Config

Config.set("graphics", "resizable", "0")
Config.set("graphics", "width", "420")
Config.set("graphics", "height", "840")

from kivymd.app import MDApp
from kivy.lang import Builder
from kivy.uix.screenmanager import Screen
from kivy.clock import Clock
from kivy.animation import Animation
from kivy.graphics.texture import Texture
from kivy.properties import NumericProperty, DictProperty, StringProperty
from kivy.utils import platform
from kivy.core.window import Window
from kivymd.uix.filemanager import MDFileManager
from kivymd.uix.boxlayout import MDBoxLayout
from kivymd.uix.button import MDFlatButton, MDFillRoundFlatIconButton
from kivymd.uix.card import MDCard
from kivymd.uix.label import MDLabel
from kivymd.uix.textfield import MDTextField
from kivymd.uix.toolbar import MDTopAppBar
from kivymd.uix.button import MDIconButton
from kivymd.uix.dialog import MDDialog
import base64
import cv2
import datetime
import html
import json
import mimetypes
import os
import shutil
import threading

from processor import SurveyProcessor
from database import Database


KV = '''
ScreenManager:
    MenuScreen:
    ConfigureScreen:
    ScannerScreen:
    PreviewScreen:
    ProcessingScreen:
    SummaryScreen:
    HistoryScreen:
    SessionDetailScreen:

<MenuScreen>:
    name: 'menu'
    MDFloatLayout:
        md_bg_color: 0.97, 0.98, 1, 1
        MDBoxLayout:
            orientation: 'vertical'
            padding: "18dp"
            spacing: "16dp"

            MDBoxLayout:
                size_hint_y: None
                height: "62dp"
                spacing: "12dp"
                MDCard:
                    size_hint: None, None
                    size: "48dp", "48dp"
                    radius: [16,]
                    md_bg_color: 0.32, 0.20, 0.92, 1
                    elevation: 0
                    MDIcon:
                        icon: "lightning-bolt-outline"
                        halign: "center"
                        valign: "center"
                        theme_text_color: "Custom"
                        text_color: 1, 1, 1, 1
                        font_size: "26sp"
                MDBoxLayout:
                    orientation: 'vertical'
                    MDLabel:
                        text: "SnapStat"
                        font_style: "H5"
                        bold: True
                        theme_text_color: "Custom"
                        text_color: 0.02, 0.03, 0.08, 1
                    MDLabel:
                        text: "Survey Analysis"
                        font_style: "Caption"
                        theme_text_color: "Custom"
                        text_color: 0.30, 0.32, 0.42, 1
                MDIconButton:
                    icon: "information-outline"
                    theme_text_color: "Custom"
                    text_color: 1, 1, 1, 1
                    md_bg_color: 0.18, 0.25, 0.92, 1
                    pos_hint: {"center_y": .5}
                    on_release: app.show_error_dialog("SnapStat", "Capture or upload a shaded survey form, then review response summaries and individual responses.")

            MDCard:
                orientation: 'vertical'
                padding: "18dp"
                spacing: "10dp"
                size_hint_y: None
                height: "212dp"
                radius: [24,]
                md_bg_color: 0.31, 0.22, 0.93, 1
                ripple_behavior: True
                on_release: app.open_config('camera', True)
                MDBoxLayout:
                    size_hint_y: None
                    height: "34dp"
                    Widget:
                    MDCard:
                        size_hint: None, None
                        size: "112dp", "34dp"
                        radius: [18,]
                        md_bg_color: 1, 1, 1, 0.22
                        elevation: 0
                        MDLabel:
                            text: "Recommended"
                            halign: "center"
                            valign: "center"
                            bold: True
                            font_style: "Caption"
                            theme_text_color: "Custom"
                            text_color: 1, 1, 1, 1
                MDCard:
                    size_hint: None, None
                    size: "80dp", "80dp"
                    radius: [40,]
                    md_bg_color: 1, 1, 1, 0.18
                    elevation: 0
                    pos_hint: {"center_x": .5}
                    MDIcon:
                        icon: "camera-outline"
                        halign: "center"
                        valign: "center"
                        font_size: "42sp"
                        theme_text_color: "Custom"
                        text_color: 1, 1, 1, 1
                MDLabel:
                    text: "Capture Survey"
                    halign: "center"
                    bold: True
                    font_style: "H6"
                    theme_text_color: "Custom"
                    text_color: 1, 1, 1, 1
                MDLabel:
                    text: "Take a photo to tally shaded responses"
                    halign: "center"
                    bold: True
                    font_style: "Caption"
                    theme_text_color: "Custom"
                    text_color: 0.88, 0.86, 1, 1

            MDBoxLayout:
                spacing: "16dp"
                size_hint_y: None
                height: "152dp"
                MDCard:
                    orientation: 'vertical'
                    padding: "18dp"
                    spacing: "8dp"
                    radius: [14,]
                    md_bg_color: 1, 1, 1, 1
                    elevation: 0
                    ripple_behavior: True
                    on_release: app.open_config('upload', True)
                    MDCard:
                        size_hint: None, None
                        size: "48dp", "48dp"
                        radius: [24,]
                        md_bg_color: 0.94, 0.94, 1, 1
                        pos_hint: {"center_x": .5}
                        elevation: 0
                        MDIcon:
                            icon: "upload-outline"
                            halign: "center"
                            valign: "center"
                            font_size: "28sp"
                            theme_text_color: "Custom"
                            text_color: 0.10, 0.20, 1, 1
                    MDLabel:
                        text: "Upload"
                        halign: "center"
                        bold: True
                        font_style: "Subtitle1"
                        theme_text_color: "Custom"
                        text_color: 0.02, 0.03, 0.08, 1
                    MDLabel:
                        text: "From gallery"
                        halign: "center"
                        font_style: "Caption"
                        theme_text_color: "Custom"
                        text_color: 0.30, 0.32, 0.42, 1
                MDCard:
                    orientation: 'vertical'
                    padding: "18dp"
                    spacing: "8dp"
                    radius: [14,]
                    md_bg_color: 1, 1, 1, 1
                    elevation: 0
                    ripple_behavior: True
                    on_release: root.manager.current = 'history'
                    MDCard:
                        size_hint: None, None
                        size: "48dp", "48dp"
                        radius: [24,]
                        md_bg_color: 0.94, 0.94, 1, 1
                        pos_hint: {"center_x": .5}
                        elevation: 0
                        MDIcon:
                            icon: "history"
                            halign: "center"
                            valign: "center"
                            font_size: "28sp"
                            theme_text_color: "Custom"
                            text_color: 0.10, 0.20, 1, 1
                    MDLabel:
                        text: "History"
                        halign: "center"
                        bold: True
                        font_style: "Subtitle1"
                        theme_text_color: "Custom"
                        text_color: 0.02, 0.03, 0.08, 1
                    MDLabel:
                        text: "View past surveys"
                        halign: "center"
                        font_style: "Caption"
                        theme_text_color: "Custom"
                        text_color: 0.30, 0.32, 0.42, 1

            Widget:

            MDCard:
                orientation: 'horizontal'
                padding: "16dp"
                spacing: "8dp"
                size_hint_y: None
                height: "58dp"
                radius: [16,]
                md_bg_color: 1, 1, 1, 1
                elevation: 0
                ripple_behavior: True
                on_release: app.show_error_dialog("Survey Form", "Use your printed shaded Likert form template for best scanning results.")
                Widget:
                MDIcon:
                    icon: "download-outline"
                    size_hint: None, None
                    size: "24dp", "24dp"
                    theme_text_color: "Custom"
                    text_color: 0.10, 0.20, 1, 1
                    pos_hint: {"center_y": .5}
                MDLabel:
                    text: "Download Survey Form"
                    bold: True
                    font_style: "Button"
                    theme_text_color: "Custom"
                    text_color: 0.02, 0.03, 0.08, 1
                    size_hint_x: None
                    width: self.texture_size[0]
                Widget:

<ConfigureScreen>:
    name: 'configure'
    MDBoxLayout:
        orientation: 'vertical'
        MDTopAppBar:
            title: "Configure Survey"
            left_action_items: [["arrow-left", lambda x: app.cancel_config()]]
        ScrollView:
            MDBoxLayout:
                id: config_body
                orientation: 'vertical'
                adaptive_height: True
                padding: "16dp"
                spacing: "14dp"
        MDBoxLayout:
            padding: "16dp"
            size_hint_y: None
            height: "78dp"
            MDFillRoundFlatIconButton:
                id: proceed_button
                icon: "camera"
                text: "Proceed to Capture"
                pos_hint: {"center_x": .5, "center_y": .5}
                size_hint_x: 1
                on_release: root.proceed()

<ScannerScreen>:
    name: 'scanner'
    FloatLayout:
        Image:
            id: preview
            allow_stretch: True
            keep_ratio: False
        MDLabel:
            text: "SYSTEM SCANNING..."
            halign: "center"
            pos_hint: {"center_y": 0.9}
            theme_text_color: "Custom"
            text_color: 0, 1, 1, 1
        MDFloatingActionButton:
            icon: "target"
            md_bg_color: 0, 1, 1, 1
            pos_hint: {"center_x": .5, "y": 0.05}
            on_release: root.capture()
        MDFloatingActionButton:
            icon: root.flash_icon
            md_bg_color: 1, 1, 1, 1
            theme_text_color: "Custom"
            text_color: 0.16, 0.16, 0.22, 1
            pos_hint: {"right": .96, "y": 0.05}
            on_release: root.toggle_flash()

<PreviewScreen>:
    name: 'preview'
    MDBoxLayout:
        orientation: 'vertical'
        MDTopAppBar:
            title: "Preview Image"
            left_action_items: [["arrow-left", lambda x: app.cancel_preview()]]
        Image:
            source: app.preview_image_path
            allow_stretch: True
            keep_ratio: True
        MDBoxLayout:
            orientation: 'vertical'
            padding: "16dp"
            spacing: "10dp"
            size_hint_y: None
            height: "180dp"
            MDFillRoundFlatIconButton:
                icon: "check-circle"
                text: "SCAN THIS IMAGE"
                pos_hint: {"center_x": .5}
                size_hint_x: 1
                on_release: app.process_preview_image()
            MDRaisedButton:
                text: "RETAKE / CHOOSE DIFFERENT"
                pos_hint: {"center_x": .5}
                size_hint_x: 1
                on_release: app.retry_preview_source()
            MDFlatButton:
                text: "CANCEL"
                pos_hint: {"center_x": .5}
                on_release: app.cancel_preview()

<ProcessingScreen>:
    name: 'processing'
    MDFloatLayout:
        md_bg_color: 0.97, 0.98, 1, 1
        MDBoxLayout:
            orientation: 'vertical'
            spacing: "18dp"
            size_hint: None, None
            size: "260dp", "220dp"
            pos_hint: {"center_x": .5, "center_y": .5}
            MDSpinner:
                size_hint: None, None
                size: "54dp", "54dp"
                pos_hint: {"center_x": .5}
                active: True
            MDLabel:
                text: app.processing_title
                halign: "center"
                bold: True
                font_style: "H6"
                theme_text_color: "Custom"
                text_color: 0.02, 0.03, 0.08, 1
            MDLabel:
                text: "Please wait while SnapStat reads the shaded answers."
                halign: "center"
                font_style: "Caption"
                theme_text_color: "Custom"
                text_color: 0.30, 0.32, 0.42, 1

<SummaryScreen>:
    name: 'summary'
    MDBoxLayout:
        orientation: 'vertical'
        MDTopAppBar:
            title: "Scan Report"
        ScrollView:
            MDBoxLayout:
                orientation: 'vertical'
                adaptive_height: True
                padding: "15dp"
                spacing: "15dp"
                MDCard:
                    orientation: 'vertical'
                    padding: "15dp"
                    size_hint_y: None
                    height: "145dp"
                    radius: [15,]
                    md_bg_color: 0.1, 0.1, 0.15, 1
                    MDLabel:
                        text: "LATEST PAPER STATUS"
                        font_style: "Overline"
                        theme_text_color: "Hint"
                    MDBoxLayout:
                        MDLabel:
                            id: completion_label
                            text: "Completion: 0/30"
                            font_style: "H5"
                        MDIcon:
                            id: warning_icon
                            icon: "alert-circle"
                            theme_text_color: "Custom"
                            text_color: 1, 0, 0, 1
                    MDLabel:
                        id: session_count_label
                        text: "Images Used: 0"
                        font_style: "Caption"
                        theme_text_color: "Hint"
                MDCard:
                    orientation: 'vertical'
                    padding: "15dp"
                    adaptive_height: True
                    radius: [15,]
                    MDLabel:
                        text: "SESSION SENTIMENT DISTRIBUTION"
                        font_style: "Button"
                    TallyRow:
                        id: bar_5
                        label: "[5] Strongly Agree"
                    TallyRow:
                        id: bar_4
                        label: "[4] Agree"
                    TallyRow:
                        id: bar_3
                        label: "[3] Neutral"
                    TallyRow:
                        id: bar_2
                        label: "[2] Disagree"
                    TallyRow:
                        id: bar_1
                        label: "[1] Strongly Disagree"
                MDBoxLayout:
                    orientation: 'vertical'
                    adaptive_height: True
                    spacing: "10dp"
                    MDLabel:
                        text: "PER QUESTION TALLY"
                        font_style: "Button"
                        theme_text_color: "Custom"
                        text_color: 0.30, 0.22, 0.86, 1
                    MDBoxLayout:
                        id: question_tally_list
                        orientation: 'vertical'
                        adaptive_height: True
                        spacing: "8dp"
                MDRaisedButton:
                    text: "SCAN ANOTHER (CAMERA)"
                    pos_hint: {"center_x": .5}
                    size_hint_x: 0.9
                    on_release: app.continue_session_capture('camera')
                MDRaisedButton:
                    text: "UPLOAD ANOTHER (STORAGE)"
                    pos_hint: {"center_x": .5}
                    size_hint_x: 0.9
                    md_bg_color: 0.2, 0.2, 0.2, 1
                    on_release: app.continue_session_capture('upload')
                MDFlatButton:
                    text: "FINISH SESSION"
                    pos_hint: {"center_x": .5}
                    on_release: app.go_home()

<TallyRow@MDBoxLayout>:
    orientation: 'vertical'
    size_hint_y: None
    height: "55dp"
    padding: [0, 5]
    label: ""
    count: "0"
    progress: 0
    MDBoxLayout:
        MDLabel:
            text: root.label
            font_style: "Caption"
        MDLabel:
            text: root.count
            halign: "right"
            font_style: "Caption"
    MDProgressBar:
        value: root.progress

<HistoryScreen>:
    name: 'history'
    MDBoxLayout:
        orientation: 'vertical'
        MDTopAppBar:
            title: "Archives"
            left_action_items: [["arrow-left", lambda x: app.go_home()]]
        ScrollView:
            MDBoxLayout:
                id: history_list
                orientation: 'vertical'
                adaptive_height: True
                padding: "15dp"
                spacing: "12dp"

<SessionDetailScreen>:
    name: 'session_detail'
    MDBoxLayout:
        orientation: 'vertical'
        MDTopAppBar:
            title: app.history_title
            left_action_items: [["arrow-left", lambda x: setattr(app.root, "current", "history")]]
        ScrollView:
            MDBoxLayout:
                id: session_detail_box
                orientation: 'vertical'
                adaptive_height: True
                padding: "15dp"
                spacing: "12dp"
'''


class MenuScreen(Screen):
    pass


class ConfigureScreen(Screen):
    def on_pre_enter(self):
        self.build_config_ui()

    def build_config_ui(self):
        app = MDApp.get_running_app()
        body = self.ids.config_body
        body.clear_widgets()
        self.validation_live = False
        self.category_rows = []
        self.scale_rows = []

        body.add_widget(
            MDLabel(
                text="Configure Categories",
                font_style="H5",
                size_hint_y=None,
                height="36dp",
            )
        )
        body.add_widget(
            MDLabel(
                text="Group survey items by topic or context.",
                theme_text_color="Hint",
                size_hint_y=None,
                height="28dp",
            )
        )

        self.category_container = MDBoxLayout(
            orientation="vertical",
            adaptive_height=True,
            spacing="12dp",
        )
        body.add_widget(self.category_container)
        for category in app.pending_categories:
            self.add_category_card(category)

        body.add_widget(
            MDFlatButton(
                text="+ Add Category",
                pos_hint={"center_x": 0.5},
                on_release=lambda *_: self.add_category_card(),
            )
        )

        body.add_widget(
            MDLabel(
                text="Configure Response Scale",
                font_style="H5",
                size_hint_y=None,
                height="42dp",
            )
        )
        body.add_widget(
            MDLabel(
                text="Customize the answer text and number of scale points.",
                theme_text_color="Hint",
                size_hint_y=None,
                height="28dp",
            )
        )

        self.scale_container = MDBoxLayout(
            orientation="vertical",
            adaptive_height=True,
            spacing="12dp",
        )
        body.add_widget(self.scale_container)
        for point in app.pending_response_scale:
            self.add_scale_card(point)

        body.add_widget(
            MDFlatButton(
                text="+ Add Scale Point",
                pos_hint={"center_x": 0.5},
                on_release=lambda *_: self.add_scale_card(),
            )
        )

        if app.pending_capture_mode == "upload":
            self.ids.proceed_button.icon = "file-upload"
            self.ids.proceed_button.text = "Proceed to Upload"
        else:
            self.ids.proceed_button.icon = "camera"
            self.ids.proceed_button.text = "Proceed to Capture"
        self.validation_live = True

    def add_category_card(self, data=None):
        data = data or self.next_category_defaults()

        card = MDCard(
            orientation="vertical",
            padding="12dp",
            spacing="8dp",
            adaptive_height=True,
            radius=[15],
        )

        header = MDBoxLayout(size_hint_y=None, height="42dp")
        header.add_widget(MDLabel(text="Category", font_style="Button"))
        header.add_widget(
            MDIconButton(
                icon="trash-can-outline",
                theme_text_color="Custom",
                text_color=(1, 0.2, 0.2, 1),
                on_release=lambda *_: self.remove_category_card(card),
            )
        )
        card.add_widget(header)

        name_input = MDTextField(
            hint_text="Category Name",
            text=str(data.get("name", "")),
            mode="rectangle",
        )
        card.add_widget(name_input)

        range_box = MDBoxLayout(
            spacing="10dp",
            size_hint_y=None,
            height="72dp",
        )
        start_input = MDTextField(
            hint_text="Start",
            text=str(data.get("start", "")),
            input_filter="int",
            mode="rectangle",
        )
        end_input = MDTextField(
            hint_text="End",
            text=str(data.get("end", "")),
            input_filter="int",
            mode="rectangle",
        )
        range_box.add_widget(start_input)
        range_box.add_widget(end_input)
        card.add_widget(range_box)

        range_label = MDLabel(
            text="Range:",
            theme_text_color="Custom",
            text_color=(0.1, 0.2, 1, 1),
            size_hint_y=None,
            height="32dp",
        )
        card.add_widget(range_label)
        error_label = MDLabel(
            text="",
            theme_text_color="Error",
            size_hint_y=None,
            height="24dp",
        )
        card.add_widget(error_label)

        row = {
            "card": card,
            "name": name_input,
            "start": start_input,
            "end": end_input,
            "range": range_label,
            "error": error_label,
        }
        self.category_rows.append(row)

        name_input.bind(text=lambda *_: self.validate_live(False))
        name_input.bind(focus=lambda _, focus: self.validate_on_blur(focus))
        start_input.bind(text=lambda *_: self.on_category_range_text(row))
        start_input.bind(focus=lambda _, focus: self.validate_on_blur(focus))
        end_input.bind(text=lambda *_: self.on_category_range_text(row))
        end_input.bind(focus=lambda _, focus: self.validate_on_blur(focus))
        self.update_category_range(row)
        self.category_container.add_widget(card)

    def on_category_range_text(self, row):
        self.update_category_range(row)
        self.validate_live(False)

    def next_category_defaults(self):
        used = set()
        for row in getattr(self, "category_rows", []):
            start = self._to_int(row["start"].text)
            end = self._to_int(row["end"].text)
            if start is None or end is None or end < start:
                continue
            used.update(range(max(1, start), min(30, end) + 1))

        available = [q_num for q_num in range(1, 31) if q_num not in used]
        if not available:
            return {
                "name": f"Category {len(self.category_rows) + 1}",
                "start": 1,
                "end": 1,
            }

        start = available[0]
        end = start
        for q_num in available[1:]:
            if q_num != end + 1:
                break
            end = q_num
        return {
            "name": f"Category {len(self.category_rows) + 1}",
            "start": start,
            "end": end,
        }

    def remove_category_card(self, card):
        self.category_container.remove_widget(card)
        self.category_rows = [
            row for row in self.category_rows
            if row["card"] is not card
        ]
        self.validate_live(False)

    def update_category_range(self, row):
        self.clear_category_error(row)
        start = self._to_int(row["start"].text)
        end = self._to_int(row["end"].text)
        if start is None or end is None or start <= 0 or end < start or end > 30:
            row["range"].text = "Range: invalid"
            row["range"].text_color = (1, 0.2, 0.2, 1)
            return
        count = end - start + 1
        row["range"].text = f"Range: {start} - {end} ({count} items)"
        row["range"].text_color = (0.1, 0.2, 1, 1)

    def add_scale_card(self, data=None):
        used_values = {
            self._to_int(row["value"].text)
            for row in getattr(self, "scale_rows", [])
        }
        available_values = [
            value for value in range(1, 6)
            if value not in used_values
        ]
        if data is None and not available_values:
            if self.scale_rows:
                self.mark_scale_error(
                    self.scale_rows[-1],
                    "All scale points 1-5 are already used.",
                )
            return

        next_value = available_values[0] if available_values else 1
        data = data or {
            "value": next_value,
            "label": f"Choice {next_value}",
        }

        card = MDCard(
            orientation="vertical",
            padding="12dp",
            spacing="8dp",
            adaptive_height=True,
            radius=[15],
        )
        header = MDBoxLayout(size_hint_y=None, height="42dp")
        header.add_widget(MDLabel(text="Scale Point", font_style="Button"))
        header.add_widget(
            MDIconButton(
                icon="trash-can-outline",
                theme_text_color="Custom",
                text_color=(1, 0.2, 0.2, 1),
                on_release=lambda *_: self.remove_scale_card(card),
            )
        )
        card.add_widget(header)

        row_box = MDBoxLayout(
            spacing="10dp",
            size_hint_y=None,
            height="72dp",
        )
        value_input = MDTextField(
            hint_text="Point",
            text=str(data.get("value", "")),
            input_filter="int",
            mode="rectangle",
            size_hint_x=0.35,
        )
        label_input = MDTextField(
            hint_text="Answer Label",
            text=str(data.get("label", "")),
            mode="rectangle",
        )
        row_box.add_widget(value_input)
        row_box.add_widget(label_input)
        card.add_widget(row_box)
        error_label = MDLabel(
            text="",
            theme_text_color="Error",
            size_hint_y=None,
            height="24dp",
        )
        card.add_widget(error_label)

        row = {
            "card": card,
            "value": value_input,
            "label": label_input,
            "error": error_label,
        }
        self.scale_rows.append(row)
        value_input.bind(text=lambda *_: self.validate_live(False))
        value_input.bind(focus=lambda _, focus: self.validate_on_blur(focus))
        label_input.bind(text=lambda *_: self.validate_live(False))
        label_input.bind(focus=lambda _, focus: self.validate_on_blur(focus))
        self.scale_container.add_widget(card)

    def remove_scale_card(self, card):
        if len(self.scale_rows) <= 1:
            return
        self.scale_container.remove_widget(card)
        self.scale_rows = [
            row for row in self.scale_rows
            if row["card"] is not card
        ]
        self.validate_live(False)

    def validate_live(self, shake):
        if not getattr(self, "validation_live", False):
            return
        self.validate_configuration(shake=shake)

    def validate_on_blur(self, focus):
        if not focus:
            self.validate_live(True)

    def proceed(self):
        app = MDApp.get_running_app()
        categories, response_scale, error = self.validate_configuration(shake=True)
        if error:
            return

        app.save_configuration(categories, response_scale)
        app.proceed_from_config()

    def validate_configuration(self, shake=True):
        self.clear_all_errors()

        used_questions = {}
        for row in self.category_rows:
            name = row["name"].text.strip()
            start = self._to_int(row["start"].text)
            end = self._to_int(row["end"].text)
            if not name:
                self.mark_category_error(
                    row,
                    "Category name cannot be empty.",
                    row["name"],
                    shake=shake,
                )
                return None, None, "Category name cannot be empty."
            if start is None or end is None or not (1 <= start <= end <= 30):
                self.mark_category_error(
                    row,
                    "Range must be inside 1-30.",
                    row["start"],
                    shake=shake,
                )
                return None, None, "Category ranges must be inside 1-30."
            for q_num in range(start, end + 1):
                if q_num in used_questions:
                    self.mark_category_error(
                        row,
                        f"Question {q_num} is already used.",
                        row["start"],
                        shake=shake,
                    )
                    self.mark_category_error(
                        used_questions[q_num],
                        f"Question {q_num} is also used here.",
                        shake=shake,
                    )
                    return None, None, "Category ranges cannot overlap."
                used_questions[q_num] = row

        scale_values = []
        for row in self.scale_rows:
            value = self._to_int(row["value"].text)
            label = row["label"].text.strip()
            if value is None or not (1 <= value <= 5):
                self.mark_scale_error(
                    row,
                    "Scale point must be 1-5.",
                    row["value"],
                    shake=shake,
                )
                return None, None, "Scale points must be from 1 to 5 only."
            if not label:
                self.mark_scale_error(
                    row,
                    "Answer label cannot be empty.",
                    row["label"],
                    shake=shake,
                )
                return None, None, "Answer labels cannot be empty."
            if value in scale_values:
                self.mark_scale_error(
                    row,
                    f"Scale point {value} is already used.",
                    row["value"],
                    shake=shake,
                )
                for other_row in self.scale_rows:
                    if other_row is not row and self._to_int(other_row["value"].text) == value:
                        self.mark_scale_error(
                            other_row,
                            f"Scale point {value} is also used here.",
                            shake=shake,
                        )
                        break
                return None, None, "Scale point values cannot be repeated."
            scale_values.append(value)

        categories = self.collect_categories()
        response_scale = self.collect_response_scale()

        if not response_scale:
            return None, None, "Add at least one valid response scale point."

        values = [point["value"] for point in response_scale]
        if len(values) != len(set(values)):
            return None, None, "Scale point values cannot be repeated."

        return categories, response_scale, ""

    def show_config_error(self, message):
        return

    def clear_all_errors(self):
        for row in getattr(self, "category_rows", []):
            self.clear_category_error(row)
        for row in getattr(self, "scale_rows", []):
            self.clear_scale_error(row)

    def clear_category_error(self, row):
        row.get("error").text = ""
        for field_name in ("name", "start", "end"):
            row[field_name].error = False

    def clear_scale_error(self, row):
        row.get("error").text = ""
        row["value"].error = False
        row["label"].error = False

    def mark_category_error(self, row, message, field=None, shake=True):
        row["error"].text = message
        if field is not None:
            field.error = True
        row["range"].text_color = (1, 0.2, 0.2, 1)
        if shake:
            self.shake_widget(row["card"])

    def mark_scale_error(self, row, message, field=None, shake=True):
        row["error"].text = message
        if field is not None:
            field.error = True
        if shake:
            self.shake_widget(row["card"])

    def shake_widget(self, widget):
        original_x = widget.x
        Animation.cancel_all(widget, "x")
        animation = (
            Animation(x=original_x - 8, duration=0.04)
            + Animation(x=original_x + 8, duration=0.04)
            + Animation(x=original_x - 6, duration=0.04)
            + Animation(x=original_x + 6, duration=0.04)
            + Animation(x=original_x, duration=0.04)
        )
        animation.start(widget)

    def collect_categories(self):
        categories = []
        for row in self.category_rows:
            start = self._to_int(row["start"].text)
            end = self._to_int(row["end"].text)
            name = row["name"].text.strip()
            if not name:
                continue
            if start is None or end is None:
                continue
            if not (1 <= start <= end <= 30):
                continue
            categories.append({
                "name": name,
                "start": start,
                "end": end,
            })
        return categories

    def collect_response_scale(self):
        points = []
        used = set()
        for row in self.scale_rows:
            value = self._to_int(row["value"].text)
            label = row["label"].text.strip()
            if value is None or not (1 <= value <= 5):
                continue
            if not label or value in used:
                continue
            points.append({"value": value, "label": label})
            used.add(value)

        return sorted(points, key=lambda point: point["value"], reverse=True)

    def _to_int(self, value):
        try:
            return int(value)
        except (TypeError, ValueError):
            return None


class ScannerScreen(Screen):
    capture_obj = None
    flash_on = False
    flash_icon = StringProperty("flashlight-off")

    def on_enter(self):
        self.capture_obj = self.open_camera()
        if not self.capture_obj or not self.capture_obj.isOpened():
            MDApp.get_running_app().show_error_dialog(
                "Camera Error",
                "The camera could not be opened. Please check camera permission "
                "or try another camera source.",
            )
            return
        Clock.schedule_interval(self.update, 1.0 / 30.0)

    def on_leave(self):
        self.stop_cam()

    def open_camera(self):
        if platform == "android":
            return self._try_camera_sources([(0, None), (1, None)])

        candidates = [(0, None)]
        if platform == "win":
            candidates.extend(
                [
                    (0, getattr(cv2, "CAP_MSMF", None)),
                    (0, getattr(cv2, "CAP_DSHOW", None)),
                    (1, None),
                    (1, getattr(cv2, "CAP_MSMF", None)),
                ]
            )
        else:
            candidates.extend([(1, None)])
        return self._try_camera_sources(candidates)

    def _try_camera_sources(self, candidates):
        for index, backend in candidates:
            if backend is None:
                capture = cv2.VideoCapture(index)
            else:
                capture = cv2.VideoCapture(index, backend)
            if capture and capture.isOpened():
                return capture
            if capture:
                capture.release()
        return None

    def toggle_flash(self):
        app = MDApp.get_running_app()
        next_state = not self.flash_on
        if platform != "android":
            app.show_error_dialog(
                "Flash Unavailable",
                "Phone flash control only works on an Android device build.",
            )
            return

        if self.set_android_torch(next_state):
            self.flash_on = next_state
            self.flash_icon = "flashlight" if self.flash_on else "flashlight-off"
        else:
            app.show_error_dialog(
                "Flash Unavailable",
                "SnapStat could not control the phone flash on this device. "
                "Please check camera permission and hardware support.",
            )

    def set_android_torch(self, enabled):
        try:
            from jnius import autoclass

            PythonActivity = autoclass("org.kivy.android.PythonActivity")
            Context = autoclass("android.content.Context")
            CameraCharacteristics = autoclass(
                "android.hardware.camera2.CameraCharacteristics"
            )

            activity = PythonActivity.mActivity
            camera_manager = activity.getSystemService(Context.CAMERA_SERVICE)
            camera_id = self.back_camera_id(camera_manager, CameraCharacteristics)
            if camera_id is None:
                return False
            camera_manager.setTorchMode(camera_id, enabled)
            return True
        except Exception:
            return False

    def back_camera_id(self, camera_manager, characteristics_class):
        for camera_id in camera_manager.getCameraIdList():
            characteristics = camera_manager.getCameraCharacteristics(camera_id)
            flash_available = characteristics.get(
                characteristics_class.FLASH_INFO_AVAILABLE
            )
            lens_facing = characteristics.get(characteristics_class.LENS_FACING)
            if (
                flash_available
                and lens_facing == characteristics_class.LENS_FACING_BACK
            ):
                return camera_id
        return None

    def update(self, dt):
        if not self.capture_obj:
            return

        ret, frame = self.capture_obj.read()
        if ret:
            frame = self.prepare_camera_frame(frame)
            pts = MDApp.get_running_app().proc.find_survey_polygon(frame)
            self.draw_scanner_overlay(frame, pts)
            buf = cv2.flip(frame, 0).tobytes()
            texture = Texture.create(
                size=(frame.shape[1], frame.shape[0]),
                colorfmt='bgr',
            )
            texture.blit_buffer(buf, colorfmt='bgr', bufferfmt='ubyte')
            self.ids.preview.texture = texture

    def prepare_camera_frame(self, frame):
        if platform == "android" and frame.shape[1] > frame.shape[0]:
            return cv2.rotate(frame, cv2.ROTATE_90_CLOCKWISE)
        return frame

    def draw_scanner_overlay(self, frame, pts):
        h, w = frame.shape[:2]
        detected = pts is not None
        guide_color = (34, 197, 94) if detected else (255, 255, 255)
        accent_color = (0, 255, 255) if detected else (190, 190, 190)
        thickness = max(2, int(min(w, h) * 0.004))

        x1, y1 = int(w * 0.08), int(h * 0.08)
        x2, y2 = int(w * 0.92), int(h * 0.86)
        corner = int(min(w, h) * 0.09)

        overlay = frame.copy()
        cv2.rectangle(overlay, (x1, y1), (x2, y2), (0, 0, 0), -1)
        cv2.addWeighted(overlay, 0.18, frame, 0.82, 0, frame)

        for sx, sy in ((x1, y1), (x2, y1), (x1, y2), (x2, y2)):
            x_dir = 1 if sx == x1 else -1
            y_dir = 1 if sy == y1 else -1
            cv2.line(frame, (sx, sy), (sx + x_dir * corner, sy), guide_color, thickness)
            cv2.line(frame, (sx, sy), (sx, sy + y_dir * corner), guide_color, thickness)

        label = "Survey form detected" if detected else "Align survey inside frame"
        label_y = max(32, y1 - 18)
        cv2.putText(
            frame,
            label,
            (x1, label_y),
            cv2.FONT_HERSHEY_SIMPLEX,
            0.7,
            guide_color,
            2,
            cv2.LINE_AA,
        )

        if detected:
            cv2.polylines(frame, [pts.astype("int32")], True, accent_color, thickness)

    def capture(self):
        if not self.capture_obj:
            return

        ret, frame = self.capture_obj.read()
        if ret:
            frame = self.prepare_camera_frame(frame)
            filename = f"scan_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S_%f')}.jpg"
            app = MDApp.get_running_app()
            path = app.app_image_path(filename)
            if cv2.imwrite(path, frame):
                app.open_image_preview(path, "camera")
            else:
                app.show_error_dialog(
                    "Camera Error",
                    "The captured image could not be saved inside the app.",
                )
        else:
            MDApp.get_running_app().show_error_dialog(
                "Camera Error",
                "No image was captured. Please try again.",
            )

    def stop_cam(self):
        Clock.unschedule(self.update)
        if self.flash_on:
            self.set_android_torch(False)
            self.flash_on = False
            self.flash_icon = "flashlight-off"
        if self.capture_obj:
            self.capture_obj.release()
            self.capture_obj = None


class PreviewScreen(Screen):
    pass


class ProcessingScreen(Screen):
    pass


class SummaryScreen(Screen):
    def on_enter(self):
        self.refresh_report()

    def refresh_report(self):
        app = MDApp.get_running_app()
        app.load_active_session_from_database()

        comp = app.last_paper_completion
        self.ids.completion_label.text = f"Latest Paper: {comp}/30"
        self.ids.warning_icon.opacity = 1 if comp < 30 else 0
        self.ids.session_count_label.text = (
            f"Images Used: {app.current_session_image_count}"
        )

        grand_total = sum(app.global_sentiment_counts.values()) or 1
        for i in range(1, 6):
            count = app.global_sentiment_counts.get(i, 0)
            percentage = (count / grand_total) * 100
            target_bar = getattr(self.ids, f"bar_{i}")
            target_bar.label = f"[{i}] {app.response_label_for(i)}"
            target_bar.count = f"({count})"
            target_bar.progress = percentage

        self._refresh_question_tally(app.question_tally_counts)

    def _refresh_question_tally(self, question_tally_counts):
        from kivymd.uix.label import MDLabel
        app = MDApp.get_running_app()
        self.ids.question_tally_list.clear_widgets()

        insights = app.session_insights(question_tally_counts)
        self.ids.question_tally_list.add_widget(
            MDLabel(
                text="Summary Insights",
                size_hint_y=None,
                height="34dp",
                font_style="Button",
                theme_text_color="Custom",
                text_color=app.theme_cls.primary_color,
            )
        )
        self.ids.question_tally_list.add_widget(
            MDLabel(
                text=insights,
                size_hint_y=None,
                height="52dp",
                font_style="Caption",
            )
        )
        app.add_question_summary_cards(
            self.ids.question_tally_list,
            question_tally_counts,
            app.category_groups(),
            app.response_scale,
        )
        return

        for group in app.category_groups():
            self.ids.question_tally_list.add_widget(
                MDLabel(
                    text=group["name"],
                    size_hint_y=None,
                    height="34dp",
                    font_style="Button",
                    theme_text_color="Custom",
                    text_color=app.theme_cls.primary_color,
                )
            )
            for q_num in range(group["start"], group["end"] + 1):
                counts = question_tally_counts[q_num]
                total = sum(counts.values())
                top_text = app.top_choice_text(counts)
                self.ids.question_tally_list.add_widget(
                    MDLabel(
                        text=f"Q{q_num}  •  {total} responses  •  {top_text}",
                        size_hint_y=None,
                        height="32dp",
                        font_style="Caption",
                    )
                )
                for choice in app.active_choice_values():
                    count = counts.get(choice, 0)
                    pct = (count / total) * 100 if total else 0
                    row = MDBoxLayout(
                        orientation="horizontal",
                        spacing="8dp",
                        size_hint_y=None,
                        height="30dp",
                    )
                    row.add_widget(
                        MDLabel(
                            text=f"{choice} {app.response_label_for(choice)}",
                            size_hint_x=0.45,
                            font_style="Caption",
                        )
                    )
                    row.add_widget(MDProgressBar(value=pct))
                    row.add_widget(
                        MDLabel(
                            text=f"{count} ({pct:.0f}%)",
                            size_hint_x=0.25,
                            halign="right",
                            font_style="Caption",
                        )
                    )
                    self.ids.question_tally_list.add_widget(row)


class HistoryScreen(Screen):
    def on_enter(self):
        self.ids.history_list.clear_widgets()
        for s in Database().get_all_sessions():
            session_id = s[0]
            date_text = s[1]
            card = MDCard(
                orientation="vertical",
                padding="12dp",
                spacing="10dp",
                adaptive_height=True,
                radius=[14],
            )
            card.add_widget(
                MDLabel(
                    text=f"Session {session_id} - {date_text}",
                    size_hint_y=None,
                    height="30dp",
                    font_style="Button",
                )
            )
            actions = MDBoxLayout(
                spacing="8dp",
                size_hint_y=None,
                height="48dp",
            )
            actions.add_widget(
                MDFillRoundFlatIconButton(
                    icon="chart-box-outline",
                    text="View",
                    on_release=lambda _btn, sid=session_id: (
                        MDApp.get_running_app().open_history_session(sid)
                    ),
                )
            )
            actions.add_widget(
                MDFillRoundFlatIconButton(
                    icon="file-excel-outline",
                    text="Excel",
                    md_bg_color=(0.05, 0.55, 0.24, 1),
                    on_release=lambda _btn, sid=session_id: (
                        MDApp.get_running_app().export_session_to_excel(sid)
                    ),
                )
            )
            actions.add_widget(
                MDIconButton(
                    icon="trash-can-outline",
                    theme_text_color="Custom",
                    text_color=(0.95, 0.12, 0.12, 1),
                    on_release=lambda _btn, sid=session_id: (
                        MDApp.get_running_app().confirm_delete_session(sid)
                    ),
                )
            )
            card.add_widget(actions)
            self.ids.history_list.add_widget(card)


class SessionDetailScreen(Screen):
    def on_enter(self):
        app = MDApp.get_running_app()
        app.populate_history_detail(self.ids.session_detail_box)


class MainApp(MDApp):
    last_paper_completion = NumericProperty(0)
    global_sentiment_counts = DictProperty({1: 0, 2: 0, 3: 0, 4: 0, 5: 0})
    question_tally_counts = DictProperty({})
    history_title = "Session"
    current_session_id = None
    selected_history_session_id = None
    current_session_image_count = 0
    pending_capture_mode = "camera"
    pending_new_session = True
    preview_image_path = StringProperty("")
    preview_source_mode = "upload"
    processing_title = StringProperty("Scanning image...")
    active_dialog = None

    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        self.file_manager = MDFileManager(
            exit_manager=self.exit_manager,
            select_path=self.select_path,
            preview=True,
            ext=[
                ".png",
                ".jpg",
                ".jpeg",
                ".jfif",
                ".bmp",
                ".webp",
                ".tif",
                ".tiff",
            ],
        )

    def build(self):
        self.lock_portrait()
        self.theme_cls.theme_style = "Light"
        self.theme_cls.primary_palette = "DeepPurple"
        self.proc = SurveyProcessor()
        self.db = Database()
        self.survey_categories = self.default_categories()
        self.response_scale = self.default_response_scale()
        self.pending_categories = [dict(item) for item in self.survey_categories]
        self.pending_response_scale = [
            dict(item) for item in self.response_scale
        ]
        self.reset_session_tallies()
        return Builder.load_string(KV)

    def on_start(self):
        self.lock_portrait()

    def lock_portrait(self):
        if platform == "android":
            try:
                from jnius import autoclass

                PythonActivity = autoclass("org.kivy.android.PythonActivity")
                ActivityInfo = autoclass("android.content.pm.ActivityInfo")
                PythonActivity.mActivity.setRequestedOrientation(
                    ActivityInfo.SCREEN_ORIENTATION_PORTRAIT
                )
            except Exception:
                pass
        else:
            if Window.width > Window.height:
                Window.size = (420, 840)

    def reset_session_tallies(self):
        self.last_paper_completion = 0
        self.current_session_image_count = 0
        self.global_sentiment_counts = {
            choice: 0 for choice in self.active_choice_values()
        }
        self.question_tally_counts = {
            q_num: {choice: 0 for choice in self.active_choice_values()}
            for q_num in range(1, 31)
        }

    def default_categories(self):
        return [
            {"name": "Customer Service", "start": 1, "end": 10},
            {"name": "Product Quality", "start": 11, "end": 20},
            {"name": "General Feedback", "start": 21, "end": 30},
        ]

    def default_response_scale(self):
        return [
            {"value": 5, "label": "Strongly Agree"},
            {"value": 4, "label": "Agree"},
            {"value": 3, "label": "Neutral"},
            {"value": 2, "label": "Disagree"},
            {"value": 1, "label": "Strongly Disagree"},
        ]

    def active_choice_values(self):
        return sorted(
            [point["value"] for point in self.response_scale],
            reverse=True,
        )

    def response_label_for(self, value):
        for point in self.response_scale:
            if point["value"] == value:
                return point["label"]
        return f"Choice {value}"

    def format_choice_count(self, value, count):
        return f"{value} {self.response_label_for(value)}={count}"

    def format_history_choice_count(self, point, count):
        return f"{point['value']} {point['label']}={count}"

    def top_choice_text(self, counts, scale=None):
        scale = scale or self.response_scale
        total = sum(counts.values())
        if total == 0:
            return "No responses yet"
        top_value = max(counts, key=lambda value: counts[value])
        label = self.scale_label_for(top_value, scale)
        return f"Top: {top_value} {label}"

    def scale_label_for(self, value, scale):
        for point in scale:
            if point["value"] == value:
                return point["label"]
        return f"Choice {value}"

    def individual_answer_text(self, value, scale):
        if value <= 0:
            return "Blank / Invalid"
        return f"{value} {self.scale_label_for(value, scale)}"

    def session_insights(self, question_counts):
        total_valid = sum(
            sum(counts.values()) for counts in question_counts.values()
        )
        possible = max(self.current_session_image_count * 30, 1)
        completion = (total_valid / possible) * 100

        choice_totals = {
            choice: self.global_sentiment_counts.get(choice, 0)
            for choice in self.active_choice_values()
        }
        if any(choice_totals.values()):
            top_value = max(choice_totals, key=lambda value: choice_totals[value])
            top = f"{top_value} {self.response_label_for(top_value)}"
        else:
            top = "No dominant answer yet"

        strongest_question = self.strongest_question_text(question_counts)
        return (
            f"{self.current_session_image_count} individual responses • "
            f"{total_valid}/{possible} valid answers ({completion:.0f}% completion)\n"
            f"Most selected: {top} • {strongest_question}"
        )

    def strongest_question_text(self, question_counts):
        best_q = None
        best_total = -1
        for q_num, counts in question_counts.items():
            total = sum(counts.values())
            if total > best_total:
                best_q = q_num
                best_total = total
        if best_q is None or best_total <= 0:
            return "No question trend yet"
        return f"Most answered: Q{best_q} ({best_total})"

    def category_groups(self, categories=None):
        if categories is None:
            categories = self.survey_categories
        valid_categories = []
        covered = set()
        for category in sorted(categories, key=lambda item: item.get("start", 0)):
            start = int(category.get("start", 0))
            end = int(category.get("end", 0))
            name = str(category.get("name", "")).strip()
            if not name or not (1 <= start <= end <= 30):
                continue
            valid_categories.append({"name": name, "start": start, "end": end})
            covered.update(range(start, end + 1))

        uncovered = [q_num for q_num in range(1, 31) if q_num not in covered]
        if uncovered:
            valid_categories.extend(self._contiguous_groups("Uncategorized", uncovered))
        return valid_categories

    def _contiguous_groups(self, name, values):
        groups = []
        start = previous = values[0]
        for value in values[1:]:
            if value == previous + 1:
                previous = value
                continue
            groups.append({"name": name, "start": start, "end": previous})
            start = previous = value
        groups.append({"name": name, "start": start, "end": previous})
        return groups

    def session_config(self):
        return {
            "categories": self.survey_categories,
            "response_scale": self.response_scale,
        }

    def add_question_summary_cards(self, container, question_counts, categories, scale):
        from kivymd.uix.card import MDCard
        from kivymd.uix.label import MDLabel
        from kivymd.uix.progressbar import MDProgressBar

        for group in categories:
            container.add_widget(
                MDLabel(
                    text=group["name"],
                    size_hint_y=None,
                    height="34dp",
                    font_style="Button",
                    theme_text_color="Custom",
                    text_color=self.theme_cls.primary_color,
                )
            )
            for q_num in range(group["start"], group["end"] + 1):
                counts = question_counts[q_num]
                total = sum(counts.values())
                card = MDCard(
                    orientation="vertical",
                    padding="14dp",
                    spacing="8dp",
                    adaptive_height=True,
                    radius=[8],
                    elevation=0,
                    md_bg_color=(1, 1, 1, 1),
                )
                card.add_widget(
                    MDLabel(
                        text=f"{q_num}. Question {q_num}",
                        size_hint_y=None,
                        height="26dp",
                        bold=True,
                        font_style="Subtitle2",
                        theme_text_color="Custom",
                        text_color=(0.02, 0.03, 0.08, 1),
                    )
                )
                card.add_widget(
                    MDLabel(
                        text=(
                            f"{total} responses  |  "
                            f"{self.top_choice_text(counts, scale)}"
                        ),
                        size_hint_y=None,
                        height="24dp",
                        font_style="Caption",
                        theme_text_color="Custom",
                        text_color=(0.25, 0.27, 0.35, 1),
                    )
                )
                for point in scale:
                    value = point["value"]
                    count = counts.get(value, 0)
                    pct = (count / total) * 100 if total else 0
                    row = MDBoxLayout(
                        orientation="horizontal",
                        spacing="8dp",
                        size_hint_y=None,
                        height="34dp",
                    )
                    row.add_widget(
                        MDLabel(
                            text=f"{value}. {point['label']}",
                            size_hint_x=0.34,
                            font_style="Caption",
                            shorten=True,
                        )
                    )
                    row.add_widget(MDProgressBar(value=pct))
                    row.add_widget(
                        MDLabel(
                            text=f"{count} ({pct:.0f}%)",
                            size_hint_x=0.22,
                            halign="right",
                            font_style="Caption",
                        )
                    )
                    card.add_widget(row)
                container.add_widget(card)

    def open_config(self, mode, new_session):
        self.pending_capture_mode = mode
        self.pending_new_session = new_session
        self.pending_categories = [dict(item) for item in self.survey_categories]
        self.pending_response_scale = [
            dict(item) for item in self.response_scale
        ]
        self.root.current = "configure"

    def cancel_config(self):
        self.root.current = "menu" if self.current_session_id is None else "summary"

    def save_configuration(self, categories, response_scale):
        if categories:
            self.survey_categories = categories
        elif categories == []:
            self.survey_categories = []
        if response_scale:
            self.response_scale = response_scale
        self.global_sentiment_counts = {
            choice: self.global_sentiment_counts.get(choice, 0)
            for choice in self.active_choice_values()
        }
        self.question_tally_counts = self._normalise_question_tally(
            self.question_tally_counts
        )

    def proceed_from_config(self):
        if self.pending_new_session:
            self.current_session_id = self.db.start_session(
                datetime.datetime.now().strftime("%Y-%m-%d %H:%M")
            )
            self.reset_session_tallies()
        else:
            self.ensure_active_session()

        if self.pending_capture_mode == "camera":
            self.root.current = "scanner"
        else:
            self.file_manager_open()

    def continue_session_capture(self, mode):
        self.ensure_active_session()
        if mode == "camera":
            self.root.current = "scanner"
        else:
            self.file_manager_open()

    def file_manager_open(self):
        self.ensure_active_session()
        path = os.path.expanduser("~")
        self.file_manager.show(path)

    def app_storage_dir(self):
        base_dir = getattr(self, "user_data_dir", None) or os.getcwd()
        storage_dir = os.path.join(base_dir, "snapstat_data")
        os.makedirs(storage_dir, exist_ok=True)
        return storage_dir

    def app_image_dir(self):
        image_dir = os.path.join(self.app_storage_dir(), "images")
        os.makedirs(image_dir, exist_ok=True)
        return image_dir

    def app_export_dir(self):
        export_dir = os.path.join(
            os.path.expanduser("~"),
            "Documents",
            "SnapStat Exports",
        )
        os.makedirs(export_dir, exist_ok=True)
        return export_dir

    def app_image_path(self, filename):
        return os.path.join(self.app_image_dir(), filename)

    def copy_image_into_app(self, source_path):
        if not source_path or not os.path.exists(source_path):
            self.show_error_dialog(
                "Image Error",
                "The selected image could not be found.",
            )
            return None

        ext = os.path.splitext(source_path)[1].lower() or ".jpg"
        filename = (
            f"upload_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S_%f')}"
            f"{ext}"
        )
        destination = self.app_image_path(filename)
        try:
            shutil.copy2(source_path, destination)
        except OSError as exc:
            self.show_error_dialog(
                "Image Error",
                f"Could not save a copy inside the app:\n{exc}",
            )
            return None
        return destination

    def select_path(self, path):
        self.exit_manager()
        self.ensure_active_session()
        copied_path = self.copy_image_into_app(path)
        if copied_path:
            self.open_image_preview(copied_path, "upload")

    def exit_manager(self, *args):
        self.file_manager.close()

    def open_image_preview(self, path, source_mode):
        if not path or not os.path.exists(path):
            self.show_error_dialog(
                "Image Error",
                "The selected image could not be found.",
            )
            return
        self.preview_image_path = path
        self.preview_source_mode = source_mode
        self.root.current = "preview"

    def cancel_preview(self):
        preview_path = self.preview_image_path
        has_scanned_images = (
            self.current_session_id is not None
            and self.db.count_session_images(self.current_session_id) > 0
        )
        self.preview_image_path = ""
        self.delete_app_image_copy(preview_path)
        if has_scanned_images:
            self.show_summary_report()
        else:
            self.root.current = "menu"

    def retry_preview_source(self):
        source_mode = self.preview_source_mode
        preview_path = self.preview_image_path
        self.preview_image_path = ""
        self.delete_app_image_copy(preview_path)
        if source_mode == "camera":
            self.root.current = "scanner"
        else:
            self.file_manager_open()

    def delete_app_image_copy(self, path):
        if not path:
            return
        image_root = os.path.abspath(self.app_image_dir())
        image_path = os.path.abspath(path)
        try:
            is_app_copy = os.path.commonpath([image_root, image_path]) == image_root
        except ValueError:
            is_app_copy = False
        if is_app_copy and os.path.exists(image_path):
            try:
                os.remove(image_path)
            except OSError:
                pass

    def process_preview_image(self, allow_blurry=False):
        path = self.preview_image_path
        if not path or not os.path.exists(path):
            self.show_error_dialog(
                "Image Error",
                "The selected image could not be found.",
            )
            return

        blur_score = self.image_blur_score(path)
        if blur_score is None:
            self.show_error_dialog(
                "Image Error",
                "This image cannot be opened for checking.",
            )
            return

        if blur_score < 35 and not allow_blurry:
            self.show_blur_warning(blur_score)
            return

        self.processing_title = "Scanning image..."
        self.root.current = "processing"
        Clock.schedule_once(lambda _dt: self.start_scan_worker(path), 0.15)

    def start_scan_worker(self, path):
        worker = threading.Thread(
            target=self.scan_image_worker,
            args=(path,),
            daemon=True,
        )
        worker.start()

    def scan_image_worker(self, path):
        results = self.proc.scan(path)
        error_message = self.proc.last_error or (
            "This image could not be processed as a survey form."
        )
        Clock.schedule_once(
            lambda _dt: self.finish_scan_worker(path, results, error_message),
            0,
        )

    def finish_scan_worker(self, path, results, error_message):
        if results:
            self.update_session_data(results, path)
            self.preview_image_path = ""
            self.show_summary_report()
            return

        self.root.current = "preview"
        self.show_error_dialog("Scan Failed", error_message)

    def image_blur_score(self, path):
        img = cv2.imread(path)
        if img is None:
            return None
        gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
        return float(cv2.Laplacian(gray, cv2.CV_64F).var())

    def show_blur_warning(self, blur_score):
        self.close_dialog()
        self.active_dialog = MDDialog(
            title="Image Looks Blurry",
            text=(
                "This image may be too blurry to read accurately. "
                f"Sharpness score: {blur_score:.1f}."
            ),
            buttons=[
                MDFlatButton(
                    text="RETAKE",
                    on_release=lambda *_: self.close_dialog_and_retry(),
                ),
                MDFlatButton(
                    text="SCAN ANYWAY",
                    on_release=lambda *_: self.close_dialog_and_process_blurry(),
                ),
            ],
        )
        self.active_dialog.open()

    def show_error_dialog(self, title, message):
        self.close_dialog()
        self.active_dialog = MDDialog(
            title=title,
            text=message,
            buttons=[
                MDFlatButton(
                    text="OK",
                    on_release=lambda *_: self.close_dialog(),
                ),
            ],
        )
        self.active_dialog.open()

    def show_info_dialog(self, title, message):
        self.show_error_dialog(title, message)

    def confirm_delete_session(self, session_id):
        self.close_dialog()
        self.active_dialog = MDDialog(
            title=f"Delete Session {session_id}?",
            text=(
                "Only this archived session and the photo copies saved inside "
                "SnapStat will be deleted. Original photos on the device will "
                "not be touched."
            ),
            buttons=[
                MDFlatButton(
                    text="CANCEL",
                    on_release=lambda *_: self.close_dialog(),
                ),
                MDFlatButton(
                    text="DELETE",
                    theme_text_color="Custom",
                    text_color=(1, 0.1, 0.1, 1),
                    on_release=lambda *_: self.delete_session(session_id),
                ),
            ],
        )
        self.active_dialog.open()

    def delete_session(self, session_id):
        self.close_dialog()
        self.db.delete_session(session_id, owned_root=self.app_image_dir())
        if self.current_session_id == session_id:
            self.current_session_id = None
            self.reset_session_tallies()
        if self.selected_history_session_id == session_id:
            self.selected_history_session_id = None
        if self.root.current == "history":
            self.root.get_screen("history").on_enter()
        elif self.root.current == "session_detail":
            self.root.current = "history"
        self.show_info_dialog(
            "Session Deleted",
            f"Session {session_id} was deleted. Original device photos were kept.",
        )

    def export_session_to_excel(self, session_id):
        session = self.db.get_session(session_id)
        if not session:
            self.show_error_dialog("Export Failed", "Session not found.")
            return

        _, date_str, tally_json, question_tally_json, config_json = session
        config = self._load_config_value(config_json)
        scale = config.get("response_scale", self.default_response_scale())
        categories = config.get("categories", self.default_categories())
        global_counts = self._json_to_int_dict(tally_json, default={})
        question_counts = self._normalise_question_tally(
            self._json_to_int_dict(question_tally_json, default={}),
            scale=scale,
        )
        images = self.db.get_session_images(session_id)

        output_dir = self.app_export_dir()
        safe_date = str(date_str).replace(":", "-").replace("/", "-")
        output_path = os.path.join(
            output_dir,
            f"session_{session_id}_{safe_date}.xls",
        )

        try:
            workbook_html = self.build_excel_html(
                session_id,
                date_str,
                scale,
                categories,
                global_counts,
                question_counts,
                images,
            )
            with open(output_path, "w", encoding="utf-8") as export_file:
                export_file.write(workbook_html)
        except OSError as exc:
            self.show_error_dialog(
                "Export Failed",
                f"Could not write the Excel file:\n{exc}",
            )
            return

        self.show_info_dialog(
            "Excel Exported",
            f"Saved Excel-compatible file to:\n{output_path}",
        )

    def build_excel_html(
        self,
        session_id,
        date_str,
        scale,
        categories,
        global_counts,
        question_counts,
        images,
    ):
        rows = [
            "<html><head><meta charset='utf-8'>",
            "<style>",
            "body{font-family:Arial,sans-serif;color:#111827}",
            "h1,h2{color:#312e81}",
            "table{border-collapse:collapse;margin-bottom:24px}",
            "th{background:#4f46e5;color:white;font-weight:bold}",
            "td,th{border:1px solid #d1d5db;padding:8px;vertical-align:top}",
            ".muted{color:#6b7280}",
            ".photo{max-width:420px;max-height:320px}",
            "</style></head><body>",
            f"<h1>SnapStat Survey Session {session_id}</h1>",
            f"<p class='muted'>Date: {html.escape(str(date_str))}</p>",
            f"<p class='muted'>Images Used: {len(images)}</p>",
            "<h2>Response Scale Summary</h2>",
            "<table><tr><th>Point</th><th>Label</th><th>Total</th></tr>",
        ]

        for point in scale:
            value = point["value"]
            rows.append(
                "<tr>"
                f"<td>{value}</td>"
                f"<td>{html.escape(str(point['label']))}</td>"
                f"<td>{global_counts.get(value, 0)}</td>"
                "</tr>"
            )
        rows.append("</table>")

        rows.append("<h2>Question Tally</h2>")
        header = ["Category", "Question", "Responses"]
        header.extend(f"{point['value']} {point['label']}" for point in scale)
        header.append("Top Answer")
        rows.append(self.excel_html_header(header))
        for group in self.category_groups(categories):
            for q_num in range(group["start"], group["end"] + 1):
                counts = question_counts[q_num]
                total = sum(counts.values())
                cells = [group["name"], q_num, total]
                cells.extend(counts.get(point["value"], 0) for point in scale)
                cells.append(self.top_choice_text(counts, scale))
                rows.append(self.excel_html_row(cells))
        rows.append("</table>")

        rows.append("<h2>Individual Responses</h2>")
        header = ["Image #", "Copied Photo Path", "Valid Answers"]
        header.extend(f"Q{q_num}" for q_num in range(1, 31))
        rows.append(self.excel_html_header(header))
        for index, (img_path, result_json) in enumerate(images, start=1):
            result = self._json_to_int_dict(result_json, default={})
            valid = len([value for value in result.values() if value > 0])
            cells = [index, img_path, valid]
            for q_num in range(1, 31):
                value = result.get(q_num, 0)
                cells.append(self.individual_answer_text(value, scale))
            rows.append(self.excel_html_row(cells))
        rows.append("</table>")

        rows.append("<h2>Photos Used</h2>")
        rows.append(
            "<table><tr><th>Image #</th><th>Copied Path</th><th>Photo</th></tr>"
        )
        for index, (img_path, _result_json) in enumerate(images, start=1):
            if os.path.exists(img_path):
                image_html = (
                    f"<img class='photo' src='{self.image_data_uri(img_path)}'>"
                )
            else:
                image_html = "Image file missing"
            rows.append(
                "<tr>"
                f"<td>{index}</td>"
                f"<td>{html.escape(str(img_path))}</td>"
                f"<td>{image_html}</td>"
                "</tr>"
            )
        rows.append("</table></body></html>")
        return "\n".join(rows)

    def excel_html_header(self, headers):
        return (
            "<table><tr>"
            + "".join(f"<th>{html.escape(str(header))}</th>" for header in headers)
            + "</tr>"
        )

    def excel_html_row(self, cells):
        return (
            "<tr>"
            + "".join(f"<td>{html.escape(str(cell))}</td>" for cell in cells)
            + "</tr>"
        )

    def image_data_uri(self, path):
        mime_type = mimetypes.guess_type(path)[0] or "image/jpeg"
        with open(path, "rb") as image_file:
            encoded = base64.b64encode(image_file.read()).decode("ascii")
        return f"data:{mime_type};base64,{encoded}"

    def write_excel_summary(self, ws, session_id, date_str, scale, global_counts, images):
        from openpyxl.styles import Font, PatternFill

        ws["A1"] = "SnapStat Survey Session"
        ws["A1"].font = Font(size=18, bold=True)
        ws["A2"] = f"Session ID: {session_id}"
        ws["A3"] = f"Date: {date_str}"
        ws["A4"] = f"Images Used: {len(images)}"
        ws["A6"] = "Response Scale"
        ws["A6"].font = Font(bold=True)
        ws["A7"] = "Point"
        ws["B7"] = "Label"
        ws["C7"] = "Total"
        for col in ("A", "B", "C"):
            ws[f"{col}7"].font = Font(bold=True, color="FFFFFF")
            ws[f"{col}7"].fill = PatternFill("solid", fgColor="4F46E5")

        row = 8
        for point in scale:
            value = point["value"]
            ws.cell(row, 1).value = value
            ws.cell(row, 2).value = point["label"]
            ws.cell(row, 3).value = global_counts.get(value, 0)
            row += 1
        self.autosize_sheet(ws)

    def write_excel_question_tally(self, ws, scale, categories, question_counts):
        headers = ["Category", "Question", "Responses"]
        for point in scale:
            headers.append(f"{point['value']} {point['label']}")
        headers.append("Top Answer")
        ws.append(headers)
        self.style_header_row(ws)

        for group in self.category_groups(categories):
            for q_num in range(group["start"], group["end"] + 1):
                counts = question_counts[q_num]
                total = sum(counts.values())
                row = [group["name"], q_num, total]
                for point in scale:
                    row.append(counts.get(point["value"], 0))
                row.append(self.top_choice_text(counts, scale))
                ws.append(row)
        self.autosize_sheet(ws)

    def write_excel_individual_responses(self, ws, scale, images):
        headers = ["Image #", "Path", "Valid Answers"]
        headers.extend([f"Q{q_num}" for q_num in range(1, 31)])
        ws.append(headers)
        self.style_header_row(ws)

        for index, (img_path, result_json) in enumerate(images, start=1):
            result = self._json_to_int_dict(result_json, default={})
            valid = len([value for value in result.values() if value > 0])
            row = [index, img_path, valid]
            for q_num in range(1, 31):
                value = result.get(q_num, 0)
                row.append(self.individual_answer_text(value, scale))
            ws.append(row)
        self.autosize_sheet(ws, max_width=38)

    def write_excel_images(self, ws, images, ExcelImage):
        ws["A1"] = "Photos Used"
        ws["A1"].font = Font(size=16, bold=True)
        row = 3
        for index, (img_path, _result_json) in enumerate(images, start=1):
            ws.cell(row, 1).value = f"Image {index}"
            ws.cell(row, 2).value = img_path
            ws.row_dimensions[row].height = 95
            if os.path.exists(img_path):
                try:
                    excel_img = ExcelImage(img_path)
                    excel_img.width = 120
                    excel_img.height = 90
                    ws.add_image(excel_img, f"C{row}")
                except Exception:
                    ws.cell(row, 3).value = "Could not embed image"
            else:
                ws.cell(row, 3).value = "Image file missing"
            row += 6
        self.autosize_sheet(ws, max_width=48)

    def style_header_row(self, ws):
        from openpyxl.styles import Font, PatternFill, Alignment

        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="4F46E5")
            cell.alignment = Alignment(horizontal="center")
        ws.freeze_panes = "A2"

    def autosize_sheet(self, ws, max_width=32):
        from openpyxl.utils import get_column_letter

        for column_cells in ws.columns:
            column_letter = get_column_letter(column_cells[0].column)
            max_len = 0
            for cell in column_cells:
                max_len = max(max_len, len(str(cell.value or "")))
            ws.column_dimensions[column_letter].width = min(max_len + 2, max_width)

    def close_dialog(self):
        if self.active_dialog:
            self.active_dialog.dismiss()
            self.active_dialog = None

    def close_dialog_and_retry(self):
        self.close_dialog()
        self.retry_preview_source()

    def close_dialog_and_process_blurry(self):
        self.close_dialog()
        self.process_preview_image(allow_blurry=True)

    def update_session_data(self, results, path):
        self.ensure_active_session()
        answered = len([v for v in results.values() if v > 0])
        self.last_paper_completion = answered

        global_counts = dict(self.global_sentiment_counts)
        question_counts = {
            q_num: dict(counts)
            for q_num, counts in self.question_tally_counts.items()
        }

        for q_num, val in results.items():
            if val in global_counts and val > 0:
                global_counts[val] += 1
                question_counts[q_num][val] += 1

        self.global_sentiment_counts = global_counts
        self.question_tally_counts = question_counts
        self._save_scan_to_database(results, path)
        self.current_session_image_count = self.db.count_session_images(
            self.current_session_id
        )

    def ensure_active_session(self):
        if self.current_session_id is None:
            self.current_session_id = self.db.start_session(
                datetime.datetime.now().strftime("%Y-%m-%d %H:%M")
            )

    def _save_scan_to_database(self, results, path):
        self.db.add_scan_to_session(
            self.current_session_id,
            path,
            results,
            self.global_sentiment_counts,
            self.question_tally_counts,
            self.session_config(),
        )

    def show_summary_report(self):
        if self.root.current == "summary":
            self.root.get_screen("summary").refresh_report()
        else:
            self.root.current = "summary"

    def load_active_session_from_database(self):
        if self.current_session_id is None:
            return

        session = self.db.get_session(self.current_session_id)
        if not session:
            return

        _, _, tally_json, question_tally_json, config_json = session
        self.apply_config_from_json(config_json)
        stored_global_counts = self._json_to_int_dict(tally_json, default={})
        if stored_global_counts:
            self.global_sentiment_counts = {
                choice: stored_global_counts.get(choice, 0)
                for choice in self.active_choice_values()
            }

        stored_question_counts = self._json_to_int_dict(
            question_tally_json,
            default={},
        )
        if stored_question_counts:
            self.question_tally_counts = self._normalise_question_tally(
                stored_question_counts
            )

        self.current_session_image_count = self.db.count_session_images(
            self.current_session_id
        )

    def open_history_session(self, session_id):
        self.selected_history_session_id = session_id
        session = self.db.get_session(session_id)
        if session:
            self.history_title = f"Session {session[0]} - {session[1]}"
        self.root.current = "session_detail"

    def populate_history_detail(self, container):
        from kivy.uix.image import Image
        from kivymd.uix.label import MDLabel
        from kivymd.uix.card import MDCard
        from kivymd.uix.progressbar import MDProgressBar

        container.clear_widgets()
        session = self.db.get_session(self.selected_history_session_id)
        if not session:
            container.add_widget(MDLabel(text="Session not found."))
            return

        _, date_str, tally_json, question_tally_json, config_json = session
        history_config = self._load_config_value(config_json)
        history_scale = history_config.get(
            "response_scale",
            self.default_response_scale(),
        )
        history_categories = history_config.get(
            "categories",
            self.default_categories(),
        )
        global_counts = self._json_to_int_dict(tally_json, default={})
        question_counts = self._normalise_question_tally(
            self._json_to_int_dict(question_tally_json, default={}),
            scale=history_scale,
        )

        container.add_widget(
            MDLabel(
                text=f"Date: {date_str}",
                size_hint_y=None,
                height="32dp",
                font_style="Button",
            )
        )
        container.add_widget(
            MDLabel(
                text=(
                    "Overall: "
                    + "  ".join(
                        self.format_history_choice_count(
                            point,
                            global_counts.get(point["value"], 0),
                        )
                        for point in history_scale
                    )
                ),
                size_hint_y=None,
                height="32dp",
            )
        )
        container.add_widget(
            MDLabel(
                text="Responses Summary",
                size_hint_y=None,
                height="38dp",
                font_style="Button",
                theme_text_color="Custom",
                text_color=self.theme_cls.primary_color,
            )
        )
        self.add_question_summary_cards(
            container,
            question_counts,
            self.category_groups(history_categories),
            history_scale,
        )

        for group in []:
            container.add_widget(
                MDLabel(
                    text=group["name"],
                    size_hint_y=None,
                    height="34dp",
                    font_style="Button",
                    theme_text_color="Custom",
                    text_color=self.theme_cls.primary_color,
                )
            )
            for q_num in range(group["start"], group["end"] + 1):
                counts = question_counts[q_num]
                total = sum(counts.values())
                container.add_widget(
                    MDLabel(
                        text=(
                            f"Q{q_num}  •  {total} responses  •  "
                            f"{self.top_choice_text(counts, history_scale)}"
                        ),
                        size_hint_y=None,
                        height="32dp",
                        font_style="Caption",
                    )
                )
                for point in history_scale:
                    count = counts[point["value"]]
                    pct = (count / total) * 100 if total else 0
                    row = MDBoxLayout(
                        orientation="horizontal",
                        spacing="8dp",
                        size_hint_y=None,
                        height="30dp",
                    )
                    row.add_widget(
                        MDLabel(
                            text=f"{point['value']} {point['label']}",
                            size_hint_x=0.45,
                            font_style="Caption",
                        )
                    )
                    row.add_widget(MDProgressBar(value=pct))
                    row.add_widget(
                        MDLabel(
                            text=f"{count} ({pct:.0f}%)",
                            size_hint_x=0.25,
                            halign="right",
                            font_style="Caption",
                        )
                    )
                    container.add_widget(row)

        container.add_widget(
            MDLabel(
                text="Response Scale: "
                + "  ".join(
                    f"{point['value']} = {point['label']}"
                    for point in history_scale
                ),
                size_hint_y=None,
                height="42dp",
                font_style="Caption",
            )
        )

        images = self.db.get_session_images(self.selected_history_session_id)
        container.add_widget(
            MDLabel(
                text=f"Individual Responses: {len(images)}",
                size_hint_y=None,
                height="36dp",
                font_style="Button",
            )
        )
        for index, (img_path, result_json) in enumerate(images, start=1):
            card = MDCard(
                orientation="vertical",
                padding="10dp",
                spacing="8dp",
                adaptive_height=True,
                radius=[15],
            )
            result = self._json_to_int_dict(result_json, default={})
            answered = len([v for v in result.values() if v > 0])
            card.add_widget(
                MDLabel(
                    text=f"Image {index}: {img_path}",
                    size_hint_y=None,
                    height="42dp",
                    font_style="Caption",
                )
            )
            if os.path.exists(img_path):
                card.add_widget(
                    Image(
                        source=img_path,
                        size_hint_y=None,
                        height="180dp",
                        allow_stretch=True,
                        keep_ratio=True,
                    )
                )
            card.add_widget(
                MDLabel(
                    text=f"Valid answers: {answered}/30",
                    size_hint_y=None,
                    height="28dp",
                    font_style="Caption",
                )
            )
            for group in self.category_groups(history_categories):
                card.add_widget(
                    MDLabel(
                        text=group["name"],
                        size_hint_y=None,
                        height="30dp",
                        font_style="Button",
                    )
                )
                for q_num in range(group["start"], group["end"] + 1):
                    value = result.get(q_num, 0)
                    answer_text = self.individual_answer_text(value, history_scale)
                    card.add_widget(
                        MDLabel(
                            text=f"Q{q_num}: {answer_text}",
                            size_hint_y=None,
                            height="24dp",
                            font_style="Caption",
                        )
                    )
            container.add_widget(card)

    def _json_to_int_dict(self, json_text, default):
        if not json_text:
            return default
        try:
            value = json.loads(json_text)
        except (TypeError, json.JSONDecodeError):
            return default

        if isinstance(value, dict):
            converted = {}
            for key, item in value.items():
                try:
                    key = int(key)
                except (TypeError, ValueError):
                    pass

                if isinstance(item, dict):
                    converted[key] = {
                        int(inner_key): inner_value
                        for inner_key, inner_value in item.items()
                    }
                else:
                    converted[key] = item
            return converted

        return default

    def _load_config_value(self, json_text):
        if not json_text:
            return {}
        try:
            value = json.loads(json_text)
        except (TypeError, json.JSONDecodeError):
            return {}
        return value if isinstance(value, dict) else {}

    def apply_config_from_json(self, json_text):
        config = self._load_config_value(json_text)
        if isinstance(config.get("categories"), list):
            self.survey_categories = config["categories"]
        if isinstance(config.get("response_scale"), list):
            self.response_scale = config["response_scale"]

    def _normalise_question_tally(self, value, scale=None):
        scale = scale or self.response_scale
        choices = sorted(
            [point["value"] for point in scale],
            reverse=True,
        )
        question_counts = {
            q_num: {choice: 0 for choice in choices}
            for q_num in range(1, 31)
        }
        for q_num, counts in value.items():
            if q_num in question_counts and isinstance(counts, dict):
                for choice in choices:
                    question_counts[q_num][choice] = counts.get(choice, 0)
        return question_counts

    def go_home(self):
        self.current_session_id = None
        self.root.current = 'menu'


if __name__ == "__main__":
    MainApp().run()
